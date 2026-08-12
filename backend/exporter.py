"""
导出模块 - 生成 Excel 质检表 + 带气泡标注的图纸图片
"""
import math
import io
from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# 类型中英映射
TYPE_CN = {
    "linear": "线性",
    "diameter": "直径",
    "radius": "半径",
    "angle": "角度",
    "chamfer": "倒角",
    "thread": "螺纹",
    "gdt": "形位公差",
    "text": "文字",
}

# 气泡配色
TYPE_COLORS = {
    "linear": "#4CAF50",
    "diameter": "#2196F3",
    "radius": "#FF9800",
    "angle": "#9C27B0",
    "chamfer": "#607D8B",
    "thread": "#F44336",
    "gdt": "#00BCD4",
    "text": "#795548",
}


def export_excel(dimensions: List[Dict], project_name: str = "质检报告") -> bytes:
    """生成 Excel 质检表，返回 bytes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "尺寸质检表"

    # 样式
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="微软雅黑", size=9)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 标题行
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{project_name} - 尺寸质检表"
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头
    headers = ["序号", "类型", "符号", "名义尺寸", "上偏差", "下偏差", "数量", "完整标注", "实测值", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 25

    # 数据行
    row = 3
    for dim in dimensions:
        if dim.get("disabled"):
            continue
        qty = dim.get("quantity", 1)
        for q in range(qty):
            num = f"{dim['id']}-{q+1}" if qty > 1 else str(dim["id"])
            full_text = dim.get("raw_text", "")

            # 公差：有则填，没有则留空（不自动猜默认公差）
            upper = dim.get("upper_tol", "") or ""
            lower = dim.get("lower_tol", "") or ""

            values = [
                num,
                TYPE_CN.get(dim.get("type", ""), dim.get("type", "")),  # 中文类型
                dim.get("symbol", ""),
                dim.get("nominal", ""),
                upper,
                lower,
                1 if qty > 1 else qty,
                full_text,
                "",  # 实测值留空
                dim.get("note", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
            row += 1

    # 列宽
    col_widths = [8, 10, 8, 12, 10, 10, 8, 25, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 底部信息
    row += 1
    ws.cell(row=row, column=1, value="检验员:").font = data_font
    ws.cell(row=row, column=3, value="__________").font = data_font
    ws.cell(row=row, column=5, value="日期:").font = data_font
    ws.cell(row=row, column=7, value="__________").font = data_font
    ws.cell(row=row, column=9, value="审核:").font = data_font
    ws.cell(row=row, column=10, value="__________").font = data_font

    # 导出为 bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _calculate_offset(cx, cy, img_w, img_h, placed_bubbles, radius):
    """
    计算气泡的避让偏移位置。
    策略：气泡偏移到尺寸标注文字的右上方，避免遮挡文字。
    如果右上方已有气泡，则尝试其他方向。
    """
    # 偏移距离 = 气泡半径 * 2.5
    offset_dist = radius * 2.5

    # 8个候选方向（优先右上，然后逆时针）
    directions = [
        (offset_dist, -offset_dist),    # 右上
        (offset_dist, 0),               # 右
        (0, -offset_dist),              # 上
        (offset_dist, offset_dist),     # 右下
        (-offset_dist, -offset_dist),   # 左上
        (-offset_dist, 0),              # 左
        (0, offset_dist),               # 下
        (-offset_dist, offset_dist),    # 左下
    ]

    for dx, dy in directions:
        new_cx = cx + dx
        new_cy = cy + dy
        # 边界检查
        if new_cx < radius or new_cx > img_w - radius:
            continue
        if new_cy < radius or new_cy > img_h - radius:
            continue
        # 碰撞检查：与其他已放置气泡的距离
        collision = False
        for px, py in placed_bubbles:
            dist = math.sqrt((new_cx - px)**2 + (new_cy - py)**2)
            if dist < radius * 2.2:  # 最小间距
                collision = True
                break
        if not collision:
            return new_cx, new_cy

    # 所有方向都有碰撞，就用右上角偏移
    new_cx = max(radius, min(cx + offset_dist, img_w - radius))
    new_cy = max(radius, min(cy - offset_dist, img_h - radius))
    return new_cx, new_cy


def export_ballooned_image(image_path: str, dimensions: List[Dict],
                            bubble_scale: float = 1.0,
                            bubble_shape: str = "circle",
                            colors: Dict[str, str] = None) -> bytes:
    """在图纸上绘制气泡标注，返回 PNG bytes"""
    img = Image.open(image_path).convert("RGBA")
    overlay = Image.new("RGBA", img.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(overlay)

    if colors is None:
        colors = TYPE_COLORS

    # 气泡尺寸放大：默认更大，导出图片清晰可辨
    base_radius = 36  # 从 18 放大到 36
    bubble_radius = int(base_radius * bubble_scale)

    # 字体也跟着放大
    font_size = int(28 * bubble_scale)
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
    except Exception:
        try:
            font = ImageFont.truetype("/opt/data/NotoSansCJKsc.otf", font_size)
        except Exception:
            font = ImageFont.load_default()

    # 引导线字体（标注文字）
    line_font_size = int(16 * bubble_scale)
    try:
        line_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", line_font_size)
    except Exception:
        line_font = ImageFont.load_default()

    placed_bubbles = []  # 已放置气泡的坐标，用于碰撞检测

    for dim in dimensions:
        if dim.get("disabled"):
            continue

        orig_cx = int(dim.get("x", 0))
        orig_cy = int(dim.get("y", 0))

        # 计算避让偏移后的位置
        cx, cy = _calculate_offset(
            orig_cx, orig_cy, img.width, img.height,
            placed_bubbles, bubble_radius
        )
        placed_bubbles.append((cx, cy))

        # 气泡颜色
        color_hex = colors.get(dim.get("type", "linear"), "#4CAF50")
        color_rgb = tuple(int(color_hex[i:i+2], 16) for i in (1, 3, 5))

        # 画引导线：从文字边缘外侧出发，避免遮挡尺寸数字
        # 计算从文字中心到气泡方向的向量，沿该方向偏移一段距离作为起点
        dx = cx - orig_cx
        dy = cy - orig_cy
        dist = math.sqrt(dx*dx + dy*dy) if (dx != 0 or dy != 0) else 1
        # 起点偏移量：沿方向远离文字中心，避开文字区域
        start_offset = 25
        line_start_x = orig_cx + int(dx / dist * start_offset)
        line_start_y = orig_cy + int(dy / dist * start_offset)
        # 终点偏移到气泡边缘，不要穿入气泡内部
        end_offset = bubble_radius - 2
        line_end_x = cx - int(dx / dist * end_offset)
        line_end_y = cy - int(dy / dist * end_offset)

        draw.line(
            [(line_start_x, line_start_y), (line_end_x, line_end_y)],
            fill=color_rgb + (200,),
            width=max(2, int(bubble_radius * 0.06)),
        )

        # 画气泡（外白底 + 内彩色层，视觉更醒目）
        if bubble_shape == "triangle":
            # 三角形：顶点在上，底边在下
            half = bubble_radius
            outer_pts = [
                (cx, cy - half),           # 顶点
                (cx - half, cy + half),    # 左下
                (cx + half, cy + half),    # 右下
            ]
            draw.polygon(outer_pts, fill=(255, 255, 255, 255), outline=color_rgb + (255,))
            # 三角形边框较粗
            draw.line(outer_pts + [outer_pts[0]], fill=color_rgb + (255,), width=max(3, int(bubble_radius * 0.12)))
            # 内彩色半透明三角形
            inner_off = int(bubble_radius * 0.18)
            inner_pts = [
                (cx, cy - half + inner_off),
                (cx - half + inner_off, cy + half - inner_off // 2),
                (cx + half - inner_off, cy + half - inner_off // 2),
            ]
            draw.polygon(inner_pts, fill=color_rgb + (180,))
        elif bubble_shape == "square":
            # 方块（圆角矩形）
            draw.rectangle(
                [cx - bubble_radius, cy - bubble_radius, cx + bubble_radius, cy + bubble_radius],
                fill=(255, 255, 255, 255),
                outline=color_rgb + (255,),
                width=max(3, int(bubble_radius * 0.12)),
            )
            # 内彩色半透明方块
            inner_r = int(bubble_radius * 0.82)
            draw.rectangle(
                [cx - inner_r, cy - inner_r, cx + inner_r, cy + inner_r],
                fill=color_rgb + (180,),
            )
        else:
            # 默认圆形
            draw.ellipse(
                [cx - bubble_radius, cy - bubble_radius, cx + bubble_radius, cy + bubble_radius],
                fill=(255, 255, 255, 255),
                outline=color_rgb + (255,),
                width=max(3, int(bubble_radius * 0.12)),
            )
            # 内彩色半透明圈
            inner_r = int(bubble_radius * 0.82)
            draw.ellipse(
                [cx - inner_r, cy - inner_r, cx + inner_r, cy + inner_r],
                fill=color_rgb + (180,),
            )

        # 编号文字
        num_text = str(dim["id"])
        bbox = draw.textbbox((0, 0), num_text, font=font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        text_x = cx - text_w // 2
        text_y = cy - text_h // 2 - bbox[1]
        draw.text((text_x, text_y), num_text, fill=(255, 255, 255, 255), font=font)

    # 合并图层
    result = Image.alpha_composite(img, overlay).convert("RGB")

    buf = io.BytesIO()
    result.save(buf, format="PNG")
    buf.seek(0)
    return buf.getvalue()
